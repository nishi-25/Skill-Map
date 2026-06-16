import csv
import io
import json
from collections import OrderedDict
from datetime import datetime

from fastapi import APIRouter, Request, Form, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response, StreamingResponse

from sqlalchemy.orm import Session

import models
import auth
from database import get_db
from template_engine import templates

router = APIRouter()


def _collect_leaf_subskill_ids(area: "models.BusinessMapArea") -> set:
    """配下（再帰的）の最下層カテゴリに割り当てられた全サブスキルIDを収集"""
    if not area.children:
        return {a_sk.sub_skill_id for a_sk in area.area_sub_skills}
    ids: set = set()
    for child in area.children:
        ids |= _collect_leaf_subskill_ids(child)
    return ids


def _build_stats_tree(areas, done_sub_ids: set) -> list:
    """エリアリストを再帰的にstats付きツリーに変換する"""
    result = []
    for area in sorted(areas, key=lambda a: a.order_index):
        sub_ids = _collect_leaf_subskill_ids(area)
        total = len(sub_ids)
        acquired = sum(1 for sid in sub_ids if sid in done_sub_ids)
        is_leaf = not area.children

        # リーフエリアのスキルデータ: エリアに含まれるスキルの全サブスキルを収集
        skills_data = []
        if is_leaf:
            seen_skills: dict = {}
            for a_sk in sorted(area.area_sub_skills, key=lambda x: x.order_index):
                ss = a_sk.sub_skill
                if ss and ss.skill and not ss.skill.is_archived:
                    skill_id = ss.skill_id
                    if skill_id not in seen_skills:
                        skill = ss.skill
                        all_subs = sorted(skill.sub_skills, key=lambda x: (x.order_index, x.id))
                        seen_skills[skill_id] = {
                            "skill": skill,
                            "sub_skills": [
                                {
                                    "id": sub.id,
                                    "name": sub.name,
                                    "description": sub.description or "",
                                    "tier": sub.tier or "basic",
                                    "can_do": sub.id in done_sub_ids,
                                }
                                for sub in all_subs
                            ],
                        }
            skills_data = list(seen_skills.values())

        result.append({
            "area": area,
            "total": total,
            "acquired": acquired,
            "pct": int(acquired / total * 100) if total else 0,
            "is_leaf": is_leaf,
            "children": _build_stats_tree(area.children, done_sub_ids) if not is_leaf else [],
            "skills_data": skills_data,
        })
    return result


# ════════════════════════════════════════════════════════════════
# ユーザー向け: スキルマップ入口の選択画面 / 業務マップ一覧
# ════════════════════════════════════════════════════════════════

@router.get("/skills/start", response_class=HTMLResponse)
def skills_start(request: Request, db: Session = Depends(get_db)):
    user = auth.require_approved(request, db)
    return templates.TemplateResponse(request, "skills_start.html", {
        "current_user": user,
    })


@router.get("/skills/business-map", response_class=HTMLResponse)
def business_map_view(request: Request, parent_id: int = 0, db: Session = Depends(get_db)):
    user = auth.require_approved(request, db)

    current_parent = None
    breadcrumbs = []
    if parent_id:
        current_parent = db.query(models.BusinessMapArea).filter(models.BusinessMapArea.id == parent_id).first()
        if current_parent is None:
            return RedirectResponse("/skills/business-map", status_code=303)
        node = current_parent
        chain = []
        while node:
            chain.append(node)
            node = node.parent
        breadcrumbs = list(reversed(chain))

    areas = (
        db.query(models.BusinessMapArea)
        .filter(models.BusinessMapArea.parent_id == (parent_id or None))
        .order_by(models.BusinessMapArea.order_index)
        .all()
    )

    # 自分が「できる」と回答したサブスキル（acquired判定用）
    done_sub_ids = {
        r.sub_skill_id
        for r in db.query(models.UserSubSkillLevel).filter(
            models.UserSubSkillLevel.user_id == user.id,
            models.UserSubSkillLevel.can_do == True,
        ).all()
    }

    tree = _build_stats_tree(areas, done_sub_ids)

    # ユーザーのエビデンスをスキルIDでインデックス化
    evidence_map: dict = {}
    for ev in (
        db.query(models.SkillEvidence)
        .filter(models.SkillEvidence.user_id == user.id)
        .order_by(models.SkillEvidence.created_at)
        .all()
    ):
        evidence_map.setdefault(ev.skill_id, []).append(ev)

    return templates.TemplateResponse(request, "business_map_view.html", {
        "current_user": user,
        "tree": tree,
        "current_parent": current_parent,
        "breadcrumbs": breadcrumbs,
        "SKILL_LEVELS": models.SKILL_LEVELS,
        "LEVEL_COLORS": models.LEVEL_COLORS,
        "SKILL_TIERS": models.SKILL_TIERS,
        "TIER_COLORS": models.TIER_COLORS,
        "evidence_map": evidence_map,
    })


@router.post("/skills/business-map/area/{area_id}/declare")
async def business_map_area_declare(area_id: int, request: Request, db: Session = Depends(get_db)):
    """業務マップのリーフエリアからサブスキルを一括申告する"""
    user = auth.require_approved(request, db)
    area = db.query(models.BusinessMapArea).filter(models.BusinessMapArea.id == area_id).first()
    if not area:
        return RedirectResponse("/skills/business-map", status_code=303)

    form_data = await request.form()
    next_url = str(form_data.get("_next", "/skills/business-map"))
    if not next_url.startswith("/"):
        next_url = "/skills/business-map"

    for a_sk in area.area_sub_skills:
        if not a_sk.sub_skill:
            continue
        ss_id = a_sk.sub_skill_id
        can_do = form_data.get(f"ss_{ss_id}") == "1"
        existing = (
            db.query(models.UserSubSkillLevel)
            .filter(
                models.UserSubSkillLevel.user_id == user.id,
                models.UserSubSkillLevel.sub_skill_id == ss_id,
            )
            .first()
        )
        if existing:
            existing.can_do = can_do
        else:
            db.add(models.UserSubSkillLevel(user_id=user.id, sub_skill_id=ss_id, can_do=can_do))

    db.commit()
    return RedirectResponse(next_url, status_code=303)


# ════════════════════════════════════════════════════════════════
# 管理: 業務マップ管理（Manager / Admin のみ）
# ════════════════════════════════════════════════════════════════

@router.get("/business-map/manage", response_class=HTMLResponse)
def business_map_manage(request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)

    areas = (
        db.query(models.BusinessMapArea)
        .filter(models.BusinessMapArea.parent_id.is_(None))
        .order_by(models.BusinessMapArea.order_index)
        .all()
    )

    # サブスキルプール: カテゴリ別にグルーピング（アーカイブ済みスキル・サブスキル無しスキルは除外）
    skills = (
        db.query(models.Skill)
        .filter(models.Skill.is_archived == False)
        .order_by(models.Skill.name)
        .all()
    )
    skills_by_category: "OrderedDict" = OrderedDict()
    for sk in skills:
        if not sk.sub_skills:
            continue
        cat_name = sk.category.name if sk.category else "未分類"
        if cat_name not in skills_by_category:
            skills_by_category[cat_name] = {
                "category": sk.category,
                "skills": [],
            }
        # スキル単位ドロップ用: 配下のサブスキル一覧をJSONで保持
        sk.sub_skills_json = json.dumps(
            [{"id": ss.id, "name": ss.name, "tier": ss.tier} for ss in sk.sub_skills],
            ensure_ascii=False,
        )
        skills_by_category[cat_name]["skills"].append(sk)

    tier_names = models.get_tier_display_names(db)

    return templates.TemplateResponse(request, "business_map_manage.html", {
        "current_user": user,
        "areas": areas,
        "skills_by_category": skills_by_category,
        "tier_names": tier_names,
    })


@router.post("/business-map/areas/new")
def business_map_area_new(
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    color: str = Form("#6366f1"),
    parent_id: int = Form(0),
    db: Session = Depends(get_db),
):
    user = auth.require_manager_or_admin(request, db)

    parent = None
    if parent_id:
        parent = db.query(models.BusinessMapArea).filter(models.BusinessMapArea.id == parent_id).first()
        if not parent:
            return RedirectResponse("/business-map/manage", status_code=303)

    parent_id_value = parent.id if parent else None
    max_order = (
        db.query(models.BusinessMapArea)
        .filter(models.BusinessMapArea.parent_id == parent_id_value)
        .count()
    )
    db.add(models.BusinessMapArea(
        name=name, description=description or None, color=color,
        order_index=max_order, parent_id=parent_id_value, created_by=user.id,
    ))
    db.commit()
    return RedirectResponse("/business-map/manage", status_code=303)


@router.post("/business-map/areas/{area_id}/edit")
def business_map_area_edit(
    area_id: int,
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    color: str = Form("#6366f1"),
    db: Session = Depends(get_db),
):
    auth.require_manager_or_admin(request, db)
    area = db.query(models.BusinessMapArea).filter(models.BusinessMapArea.id == area_id).first()
    if area:
        area.name = name
        area.description = description or None
        area.color = color
        db.commit()
    return RedirectResponse("/business-map/manage", status_code=303)


@router.post("/business-map/areas/{area_id}/delete")
def business_map_area_delete(area_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_manager_or_admin(request, db)
    area = db.query(models.BusinessMapArea).filter(models.BusinessMapArea.id == area_id).first()
    if area:
        db.delete(area)
        db.commit()
    return RedirectResponse("/business-map/manage", status_code=303)


@router.post("/business-map/areas/reorder")
async def business_map_areas_reorder(request: Request, db: Session = Depends(get_db)):
    auth.require_manager_or_admin(request, db)
    data = await request.json()
    for i, area_id in enumerate(data.get("ids", [])):
        area = db.query(models.BusinessMapArea).filter(models.BusinessMapArea.id == area_id).first()
        if area:
            area.order_index = i
    db.commit()
    return {"ok": True}


@router.post("/business-map/areas/{area_id}/skills/add")
async def business_map_area_skill_add(area_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_manager_or_admin(request, db)
    data = await request.json()
    sub_skill_id = data.get("sub_skill_id")

    area = db.query(models.BusinessMapArea).filter(models.BusinessMapArea.id == area_id).first()
    sub_skill = db.query(models.SubSkill).filter(models.SubSkill.id == sub_skill_id).first()
    if not area or not sub_skill:
        return {"ok": False}

    existing = (
        db.query(models.BusinessMapAreaSkill)
        .filter(
            models.BusinessMapAreaSkill.area_id == area_id,
            models.BusinessMapAreaSkill.sub_skill_id == sub_skill_id,
        )
        .first()
    )
    if existing:
        return {"ok": True}

    max_order = (
        db.query(models.BusinessMapAreaSkill)
        .filter(models.BusinessMapAreaSkill.area_id == area_id)
        .count()
    )
    db.add(models.BusinessMapAreaSkill(area_id=area_id, sub_skill_id=sub_skill_id, order_index=max_order))
    db.commit()
    return {"ok": True}


@router.post("/business-map/areas/{area_id}/skills/remove")
async def business_map_area_skill_remove(area_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_manager_or_admin(request, db)
    data = await request.json()
    sub_skill_id = data.get("sub_skill_id")

    db.query(models.BusinessMapAreaSkill).filter(
        models.BusinessMapAreaSkill.area_id == area_id,
        models.BusinessMapAreaSkill.sub_skill_id == sub_skill_id,
    ).delete()
    db.commit()
    return {"ok": True}


@router.post("/business-map/area-skills/reorder")
async def business_map_area_skills_reorder(request: Request, db: Session = Depends(get_db)):
    auth.require_manager_or_admin(request, db)
    data = await request.json()
    area_id = data.get("area_id")
    for i, sub_skill_id in enumerate(data.get("sub_skill_ids", [])):
        a_sk = (
            db.query(models.BusinessMapAreaSkill)
            .filter(
                models.BusinessMapAreaSkill.area_id == area_id,
                models.BusinessMapAreaSkill.sub_skill_id == sub_skill_id,
            )
            .first()
        )
        if a_sk:
            a_sk.order_index = i
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════
# 業務マップ: エクスポート / インポート
# ════════════════════════════════════════════════════════════════

# id / parent_id はファイル内でカテゴリの親子関係を表すための参照番号
# （DB上の実際のIDではない。新規インポート時は自由に振ってよい）
_BM_AREA_FIELDS = ["id", "parent_id", "name", "description", "color", "order_index"]
_BM_AREA_SKILL_FIELDS = ["area_id", "skill_name", "sub_skill_name", "order_index"]
_BM_FIELD_MAP = {
    "business_map_areas": _BM_AREA_FIELDS,
    "business_map_area_skills": _BM_AREA_SKILL_FIELDS,
}
_BM_INT_FIELDS = {"id", "parent_id", "area_id", "order_index"}

# インポートのテンプレート / 入力例として表示するサンプルデータ
_BM_TEMPLATE_DATA = {
    "business_map_areas": [
        {"id": 1, "parent_id": "", "name": "単体HILS構築", "description": "単体HILS環境の構築に関する業務", "color": "#6366f1", "order_index": 1},
        {"id": 2, "parent_id": 1, "name": "ハード構築", "description": "", "color": "#6366f1", "order_index": 1},
        {"id": 3, "parent_id": 2, "name": "Configuration設定", "description": "", "color": "#6366f1", "order_index": 1},
    ],
    "business_map_area_skills": [
        {"area_id": 3, "skill_name": "Python", "sub_skill_name": "変数とデータ型の基本操作", "order_index": 1},
    ],
}


def _bm_xlsx_bytes(data: dict) -> bytes:
    """業務マップのデータをExcelワークブックのbytesに変換"""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for key, fields in _BM_FIELD_MAP.items():
        if key not in data:
            continue
        ws = wb.create_sheet(title=key)
        ws.append(fields)
        for row in data[key]:
            ws.append([row.get(f, "") for f in fields])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bm_csv_bytes(data: dict, key: str) -> bytes:
    """業務マップのうち1エンティティ分をCSVのbytesに変換"""
    fields = _BM_FIELD_MAP[key]
    out = io.StringIO()
    out.write("﻿")
    writer = csv.writer(out)
    writer.writerow(fields)
    for row in data.get(key, []):
        writer.writerow([row.get(f, "") for f in fields])
    return out.getvalue().encode("utf-8")


def _bm_zip_bytes(data: dict, keys: list) -> bytes:
    """複数エンティティを `{エンティティ名}.csv` にまとめたZIPのbytesに変換"""
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for key in keys:
            zf.writestr(f"{key}.csv", _bm_csv_bytes(data, key))
    return buf.getvalue()


@router.get("/business-map/export")
def business_map_export(request: Request, format: str = "json", db: Session = Depends(get_db)):
    """業務マップ（カテゴリ階層・サブスキル割り当て）を一括エクスポート"""
    auth.require_manager_or_admin(request, db)

    areas = db.query(models.BusinessMapArea).order_by(models.BusinessMapArea.id).all()
    data: dict = {
        "exported_at": datetime.now().isoformat(),
        "business_map_areas": [
            {
                "id": a.id,
                "parent_id": a.parent_id or "",
                "name": a.name,
                "description": a.description or "",
                "color": a.color,
                "order_index": a.order_index,
            }
            for a in areas
        ],
        "business_map_area_skills": [
            {
                "area_id": a.id,
                "skill_name": a_sk.sub_skill.skill.name if a_sk.sub_skill and a_sk.sub_skill.skill else "",
                "sub_skill_name": a_sk.sub_skill.name if a_sk.sub_skill else "",
                "order_index": a_sk.order_index,
            }
            for a in areas
            for a_sk in a.area_sub_skills
        ],
    }

    date_str = datetime.now().strftime("%Y%m%d")

    if format == "xlsx":
        buf = io.BytesIO(_bm_xlsx_bytes(data))
        filename = f"business_map_{date_str}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    if format == "csv":
        buf = io.BytesIO(_bm_zip_bytes(data, list(_BM_FIELD_MAP)))
        filename = f"business_map_{date_str}.zip"
        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    body = json.dumps(data, ensure_ascii=False, indent=2)
    filename = f"business_map_{date_str}.json"
    return Response(
        content=body.encode("utf-8"),
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/business-map/import-template")
def business_map_import_template(request: Request, format: str = "xlsx", db: Session = Depends(get_db)):
    """業務マップ一括インポート用のテンプレート（ヘッダー＋入力例）をダウンロード"""
    auth.require_manager_or_admin(request, db)

    date_str = datetime.now().strftime("%Y%m%d")

    if format == "csv":
        buf = io.BytesIO(_bm_zip_bytes(_BM_TEMPLATE_DATA, list(_BM_FIELD_MAP)))
        filename = f"business_map_template_{date_str}.zip"
        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    buf = io.BytesIO(_bm_xlsx_bytes(_BM_TEMPLATE_DATA))
    filename = f"business_map_template_{date_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _bm_row_value(field: str, raw):
    """CSV/Excelの1セルを一括インポート用のdictの値に変換する"""
    if raw is None:
        raw = ""
    if isinstance(raw, str):
        raw = raw.strip()
    if field in _BM_INT_FIELDS:
        if raw == "":
            return None
        try:
            return int(raw)
        except (ValueError, TypeError):
            return None
    return raw


def _bm_rows_to_items(headers: list, rows) -> list:
    headers = [(h or "").strip() for h in headers]
    items = []
    for row in rows:
        if all((c is None or str(c).strip() == "") for c in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            item[h] = _bm_row_value(h, v)
        items.append(item)
    return items


def _parse_bm_xlsx(content: bytes) -> dict:
    """業務マップ一括エクスポートExcelファイル（シート名=business_map_areas/business_map_area_skills）を辞書に変換"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    data: dict = {}
    for key in _BM_FIELD_MAP:
        if key not in wb.sheetnames:
            continue
        rows = list(wb[key].iter_rows(values_only=True))
        if not rows:
            continue
        data[key] = _bm_rows_to_items(list(rows[0]), rows[1:])
    return data


def _parse_bm_csv_rows(content: bytes):
    reader = csv.reader(io.StringIO(content.decode("utf-8-sig")))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _detect_bm_csv_key(headers: list):
    hs = {(h or "").strip() for h in headers}
    if "skill_name" in hs and "sub_skill_name" in hs:
        return "business_map_area_skills"
    if "parent_id" in hs and "color" in hs:
        return "business_map_areas"
    return None


def _parse_bm_csv(content: bytes, filename: str = "") -> dict:
    """単一エンティティ分の業務マップ一括エクスポートCSVファイルを辞書に変換"""
    headers, rows = _parse_bm_csv_rows(content)
    key = _detect_bm_csv_key(headers)
    if not key:
        filename = filename.lower()
        for cand in _BM_FIELD_MAP:
            if cand in filename:
                key = cand
                break
    if not key:
        return {}
    return {key: _bm_rows_to_items(headers, rows)}


def _parse_bm_zip(content: bytes) -> dict:
    """複数CSV（business_map_areas.csv / business_map_area_skills.csv）を含むZIPを辞書に変換"""
    import zipfile
    data: dict = {}
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        for name in zf.namelist():
            base = name.rsplit("/", 1)[-1]
            for key in _BM_FIELD_MAP:
                if base.lower().startswith(key):
                    headers, rows = _parse_bm_csv_rows(zf.read(name))
                    if headers:
                        data[key] = _bm_rows_to_items(headers, rows)
                    break
    return data


def _delete_all_business_map_data(db: Session):
    """業務マップのカテゴリ階層・サブスキル割り当てを完全に削除する（「全削除してインポート」の前処理）"""
    db.query(models.BusinessMapAreaSkill).delete(synchronize_session=False)
    db.query(models.BusinessMapArea).delete(synchronize_session=False)
    db.commit()


def _apply_bm_import(data: dict, db: Session, mode: str = "add") -> dict:
    """一括エクスポート形式のデータ（業務マップのカテゴリ階層・サブスキル割り当て）をDBに反映する

    mode:
      "add"         同名・同階層のカテゴリは新規作成せずスキップする（新規追加のみ行う・デフォルト）
      "replace_all" インポート前に既存の業務マップをすべて削除してからインポートする
    """
    if mode == "replace_all":
        _delete_all_business_map_data(db)

    added_areas = skipped_areas = 0
    added_area_skills = skipped_area_skills = 0

    def process_area_row(row, parent_new_id):
        nonlocal added_areas, skipped_areas
        name = row["name"].strip()
        existing = None
        if mode == "add":
            existing = db.query(models.BusinessMapArea).filter(
                models.BusinessMapArea.name == name,
                models.BusinessMapArea.parent_id == parent_new_id,
            ).first()
        if existing:
            skipped_areas += 1
            return existing.id
        max_order = (
            db.query(models.BusinessMapArea)
            .filter(models.BusinessMapArea.parent_id == parent_new_id)
            .count()
        )
        order_index = row.get("order_index")
        new_area = models.BusinessMapArea(
            name=name,
            description=(row.get("description") or "").strip() or None,
            color=(row.get("color") or "#6366f1").strip() or "#6366f1",
            order_index=order_index if order_index is not None else max_order,
            parent_id=parent_new_id,
        )
        db.add(new_area)
        db.flush()
        added_areas += 1
        return new_area.id

    # ─ カテゴリ階層のインポート（親が先に解決済みの行から順に処理） ─
    id_map: dict = {}
    remaining = [r for r in data.get("business_map_areas", []) if (r.get("name") or "").strip()]
    pending_ids = {r.get("id") for r in remaining if r.get("id") is not None}
    while remaining:
        next_remaining = []
        progressed = False
        for row in remaining:
            old_id = row.get("id")
            parent_old_id = row.get("parent_id")
            if parent_old_id is not None and parent_old_id in pending_ids and parent_old_id not in id_map:
                next_remaining.append(row)
                continue
            parent_new_id = id_map.get(parent_old_id) if parent_old_id is not None else None
            new_id = process_area_row(row, parent_new_id)
            if old_id is not None:
                id_map[old_id] = new_id
            progressed = True
        if not progressed:
            # 循環参照など解決不能な行はルートカテゴリとして扱う
            for row in remaining:
                old_id = row.get("id")
                new_id = process_area_row(row, None)
                if old_id is not None:
                    id_map[old_id] = new_id
            break
        remaining = next_remaining

    # ─ サブスキル割り当てのインポート ─
    for row in data.get("business_map_area_skills", []):
        area_new_id = id_map.get(row.get("area_id"))
        skill_name = (row.get("skill_name") or "").strip()
        sub_skill_name = (row.get("sub_skill_name") or "").strip()
        if not area_new_id or not skill_name or not sub_skill_name:
            skipped_area_skills += 1
            continue
        sub_skill = (
            db.query(models.SubSkill)
            .join(models.Skill, models.SubSkill.skill_id == models.Skill.id)
            .filter(models.Skill.name == skill_name, models.SubSkill.name == sub_skill_name)
            .first()
        )
        if not sub_skill:
            skipped_area_skills += 1
            continue
        existing = db.query(models.BusinessMapAreaSkill).filter(
            models.BusinessMapAreaSkill.area_id == area_new_id,
            models.BusinessMapAreaSkill.sub_skill_id == sub_skill.id,
        ).first()
        if existing:
            skipped_area_skills += 1
            continue
        order_index = row.get("order_index")
        if order_index is None:
            order_index = (
                db.query(models.BusinessMapAreaSkill)
                .filter(models.BusinessMapAreaSkill.area_id == area_new_id)
                .count()
            )
        db.add(models.BusinessMapAreaSkill(
            area_id=area_new_id, sub_skill_id=sub_skill.id, order_index=order_index,
        ))
        added_area_skills += 1

    db.commit()
    return {
        "added_areas": added_areas,
        "skipped_areas": skipped_areas,
        "added_area_skills": added_area_skills,
        "skipped_area_skills": skipped_area_skills,
    }


@router.post("/business-map/import")
async def business_map_import(
    request: Request,
    file: UploadFile = File(...),
    mode: str = Form(default="add"),
    db: Session = Depends(get_db),
):
    """業務マップ一括エクスポートファイル（JSON/Excel/CSV/ZIP）からカテゴリ階層・サブスキル割り当てを一括インポート

    mode="add": 既存データに新規分のみ追加（デフォルト）
    mode="replace_all": 既存の業務マップを全削除してからインポート
    """
    auth.require_manager_or_admin(request, db)

    if mode not in ("add", "replace_all"):
        mode = "add"

    content = await file.read()
    filename = (file.filename or "").lower()

    try:
        if filename.endswith(".xlsx"):
            data = _parse_bm_xlsx(content)
        elif filename.endswith(".zip"):
            data = _parse_bm_zip(content)
        elif filename.endswith(".csv"):
            data = _parse_bm_csv(content, filename)
        else:
            data = json.loads(content.decode("utf-8-sig"))
    except Exception:
        return JSONResponse({"ok": False, "error": "ファイルの解析に失敗しました"}, status_code=400)

    result = _apply_bm_import(data, db, mode=mode)
    return JSONResponse({"ok": True, "mode": mode, **result})
