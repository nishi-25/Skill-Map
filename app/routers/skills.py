from typing import List, Optional
import csv
import io
from fastapi import APIRouter, Request, Form, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy.sql import func
from collections import defaultdict

import models
import auth
from database import get_db
from template_engine import templates
from routers.groups import _get_all_group_skill_ids

router = APIRouter()


# ════════════════════════════════════════════════════════════════
# カテゴリー管理（Admin / Manager のみ）
# ════════════════════════════════════════════════════════════════

@router.get("/categories", response_class=HTMLResponse)
def categories_list(request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)

    if user.role == "manager":
        # Managerは自グループのスキルが属するカテゴリのみ表示
        skill_ids = _get_manager_skill_ids(user, db)
        if skill_ids:
            cat_ids = {s.category_id for s in
                       db.query(models.Skill).filter(
                           models.Skill.id.in_(skill_ids),
                           models.Skill.category_id.isnot(None)
                       ).all()}
            cats = db.query(models.Category).filter(
                models.Category.id.in_(cat_ids)
            ).order_by(models.Category.name).all()
        else:
            cats = []
    else:
        cats = db.query(models.Category).order_by(models.Category.name).all()

    return templates.TemplateResponse(request, "categories.html", {
        "current_user": user, "categories": cats
    })


@router.get("/categories/new", response_class=HTMLResponse)
def category_new_get(request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)
    return templates.TemplateResponse(request, "category_form.html", {
        "current_user": user, "category": None, "error": None
    })


@router.post("/categories/new")
def category_new_post(
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    color: str = Form("#6366f1"),
    db: Session = Depends(get_db),
):
    user = auth.require_manager_or_admin(request, db)
    if db.query(models.Category).filter(models.Category.name == name).first():
        return templates.TemplateResponse(request, "category_form.html", {
            "current_user": user, "category": None,
            "error": "そのカテゴリー名は既に使用されています"
        })
    db.add(models.Category(
        name=name, description=description or None,
        color=color, created_by=user.id
    ))
    db.commit()
    return RedirectResponse("/categories", status_code=303)


@router.get("/categories/{cat_id}/edit", response_class=HTMLResponse)
def category_edit_get(cat_id: int, request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)
    cat = db.query(models.Category).filter(models.Category.id == cat_id).first()
    if not cat:
        return RedirectResponse("/categories", status_code=303)
    return templates.TemplateResponse(request, "category_form.html", {
        "current_user": user, "category": cat, "error": None
    })


@router.post("/categories/{cat_id}/edit")
def category_edit_post(
    cat_id: int,
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    color: str = Form("#6366f1"),
    db: Session = Depends(get_db),
):
    user = auth.require_manager_or_admin(request, db)
    cat = db.query(models.Category).filter(models.Category.id == cat_id).first()
    if not cat:
        return RedirectResponse("/categories", status_code=303)
    dup = db.query(models.Category).filter(
        models.Category.name == name, models.Category.id != cat_id
    ).first()
    if dup:
        return templates.TemplateResponse(request, "category_form.html", {
            "current_user": user, "category": cat,
            "error": "そのカテゴリー名は既に使用されています"
        })
    cat.name = name
    cat.description = description or None
    cat.color = color
    db.commit()
    return RedirectResponse("/categories", status_code=303)


@router.post("/categories/{cat_id}/delete")
def category_delete(cat_id: int, request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)
    cat = db.query(models.Category).filter(models.Category.id == cat_id).first()
    if cat:
        db.query(models.Skill).filter(models.Skill.category_id == cat_id).update(
            {"category_id": None}
        )
        db.delete(cat)
        db.commit()
    return RedirectResponse("/categories", status_code=303)


@router.post("/categories/bulk-delete")
def categories_bulk_delete(
    request: Request,
    cat_ids: List[int] = Form(default=[]),
    db: Session = Depends(get_db),
):
    """複数カテゴリーを一括削除（紐づくスキルは未分類になる）"""
    user = auth.require_manager_or_admin(request, db)
    if cat_ids:
        db.query(models.Skill).filter(models.Skill.category_id.in_(cat_ids)).update(
            {"category_id": None}, synchronize_session=False
        )
        db.query(models.Category).filter(models.Category.id.in_(cat_ids)).delete(
            synchronize_session=False
        )
        db.commit()
    return RedirectResponse("/categories", status_code=303)


# ════════════════════════════════════════════════════════════════
# カテゴリー エクスポート / インポート
# ════════════════════════════════════════════════════════════════

_CAT_SAMPLE = [
    {"name": "HILS基盤・操作",         "color": "#dc2626", "description": ""},
    {"name": "dSPACEツール",           "color": "#2563eb", "description": ""},
    {"name": "ソフト検証・テスト",      "color": "#7c3aed", "description": ""},
    {"name": "モデル開発（Simulink）",  "color": "#f97316", "description": ""},
]


@router.get("/categories/export/{fmt}")
def categories_export(fmt: str, request: Request, db: Session = Depends(get_db)):
    """カテゴリーを CSV / Excel / JSON / Markdown でエクスポート"""
    from fastapi.responses import StreamingResponse, Response
    user = auth.require_manager_or_admin(request, db)
    cats = db.query(models.Category).order_by(models.Category.name).all()

    if fmt == "csv":
        lines = ["name,color,description"]
        for c in cats:
            desc = (c.description or "").replace(",", "、")
            lines.append(f'{c.name},{c.color},{desc}')
        body = "﻿" + "\n".join(lines) + "\n"
        return Response(content=body.encode("utf-8"),
                        media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=categories.csv"})

    elif fmt == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "カテゴリー"
        headers = ["name", "color", "description"]
        headers_ja = ["カテゴリー名", "カラー(16進)", "説明"]
        hfill = PatternFill("solid", fgColor="F97316")
        hfont = Font(bold=True, color="FFFFFF")
        for ci, (h, hj) in enumerate(zip(headers, headers_ja), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hfill_ = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=ci, value=f"({hj})")
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 36
        for ri, c in enumerate(cats, start=3):
            ws.cell(row=ri, column=1, value=c.name)
            ws.cell(row=ri, column=2, value=c.color)
            ws.cell(row=ri, column=3, value=c.description or "")
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf,
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=categories.xlsx"})

    elif fmt == "json":
        data = [{"name": c.name, "color": c.color, "description": c.description or ""} for c in cats]
        body = _json.dumps(data, ensure_ascii=False, indent=2)
        return Response(content=body.encode("utf-8"),
                        media_type="application/json; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=categories.json"})

    elif fmt == "markdown":
        lines = ["# カテゴリー一覧\n",
                 "<!-- 書式: | name | color(16進) | description | -->",
                 "",
                 "| カテゴリー名 | カラー | 説明 |",
                 "|------------|--------|------|"]
        for c in cats:
            lines.append(f"| {c.name} | {c.color} | {c.description or ''} |")
        lines.append("")
        body = "\n".join(lines)
        return Response(content=body.encode("utf-8"),
                        media_type="text/markdown; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=categories.md"})

    else:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail=f"不明な形式: {fmt}")


@router.post("/categories/import")
async def categories_import(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """CSV / Excel / JSON / Markdown からカテゴリーを一括インポート"""
    user = auth.require_manager_or_admin(request, db)
    content  = await file.read()
    filename = (file.filename or "").lower()

    records = []
    errors  = []

    if filename.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp932", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for i, row in enumerate(reader, start=2):
            name  = (row.get("name") or row.get("カテゴリー名") or "").strip()
            color = (row.get("color") or row.get("カラー") or "#f97316").strip()
            desc  = (row.get("description") or row.get("説明") or "").strip()
            if not name: errors.append(f"行{i}: カテゴリー名が空です"); continue
            records.append({"name": name, "color": color, "description": desc})

    elif filename.endswith((".xlsx", ".xls")):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            return JSONResponse({"ok": False, "error": f"Excelファイルの読み込みに失敗: {e}"}, status_code=400)
        ws = wb["カテゴリー"] if "カテゴリー" in wb.sheetnames else wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}
        def gv(row, *keys):
            for k in keys:
                if k in col and row[col[k]].value is not None:
                    return str(row[col[k]].value).strip()
            return ""
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
            row = list(row)
            if not any(c.value for c in row): continue
            name  = gv(row, "name", "カテゴリー名")
            color = gv(row, "color", "カラー") or "#f97316"
            desc  = gv(row, "description", "説明")
            if not name: errors.append(f"行{i}: カテゴリー名が空です"); continue
            records.append({"name": name, "color": color, "description": desc})

    elif filename.endswith(".json"):
        import json as _json2
        try:
            data = _json2.loads(content.decode("utf-8-sig"))
        except Exception as e:
            return JSONResponse({"ok": False, "error": f"JSON解析エラー: {e}"}, status_code=400)
        if not isinstance(data, list):
            return JSONResponse({"ok": False, "error": "JSONはリスト形式にしてください"}, status_code=400)
        for i, item in enumerate(data):
            name  = str(item.get("name") or "").strip()
            color = str(item.get("color") or "#f97316").strip()
            desc  = str(item.get("description") or "").strip()
            if not name: errors.append(f"item[{i}]: カテゴリー名が空です"); continue
            records.append({"name": name, "color": color, "description": desc})

    elif filename.endswith((".md", ".markdown")):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp932", errors="replace")
        for line in text.splitlines():
            line = line.strip()
            if not line.startswith("|") or "カテゴリー名" in line or "---" in line: continue
            parts = [p.strip() for p in line.strip("|").split("|")]
            if len(parts) < 2: continue
            name  = parts[0].strip()
            color = parts[1].strip() if len(parts) > 1 else "#f97316"
            desc  = parts[2].strip() if len(parts) > 2 else ""
            if not name or name == "カテゴリー名": continue
            records.append({"name": name, "color": color or "#f97316", "description": desc})
    else:
        return JSONResponse({"ok": False, "error": "対応形式: .csv .xlsx .md .json"}, status_code=400)

    added = skipped = 0
    for r in records:
        existing = db.query(models.Category).filter(models.Category.name == r["name"]).first()
        if existing:
            errors.append(f"'{r['name']}' は既に存在します（スキップ）")
            skipped += 1
            continue
        db.add(models.Category(
            name=r["name"], color=r["color"],
            description=r["description"] or None,
            created_by=user.id,
        ))
        added += 1
    db.commit()
    return JSONResponse({"ok": True, "added": added, "skipped": skipped, "errors": errors})


# ════════════════════════════════════════════════════════════════
# スキルカタログ管理（Admin / Manager のみ）
# ════════════════════════════════════════════════════════════════

def _get_manager_skill_ids(user, db) -> set:
    """Managerが管理するグループに割り当てられた全スキルIDを返す"""
    from sqlalchemy import text
    rows = db.execute(
        text("SELECT DISTINCT group_id FROM group_managers WHERE user_id = :uid"),
        {"uid": user.id}
    ).fetchall()
    gm_ids = {r[0] for r in rows}
    primary_ids = {g.id for g in db.query(models.Group).filter(
        models.Group.manager_id == user.id
    ).all()}
    all_group_ids = gm_ids | primary_ids
    if not all_group_ids:
        return set()
    groups = db.query(models.Group).filter(models.Group.id.in_(all_group_ids)).all()
    skill_ids: set = set()
    for g in groups:
        skill_ids |= _get_all_group_skill_ids(g)
    return skill_ids


@router.get("/skills/catalog", response_class=HTMLResponse)
def catalog_list(
    request: Request,
    category_id: int = 0,
    tier: str = "",
    db: Session = Depends(get_db),
):
    user = auth.require_manager_or_admin(request, db)
    q = db.query(models.Skill)

    if user.role == "manager":
        # Managerは自グループに割り当てられたスキルのみ表示
        skill_ids = _get_manager_skill_ids(user, db)
        q = q.filter(models.Skill.id.in_(skill_ids)) if skill_ids else q.filter(models.Skill.id.in_([]))

    if category_id:
        q = q.filter(models.Skill.category_id == category_id)
    if tier:
        q = q.filter(models.Skill.tier == tier)
    skills = q.order_by(models.Skill.tier, models.Skill.name).all()

    # カテゴリもManagerのスコープに絞る
    if user.role == "manager":
        skill_ids_all = _get_manager_skill_ids(user, db)
        cat_ids = {s.category_id for s in skills if s.category_id}
        categories = db.query(models.Category).filter(
            models.Category.id.in_(cat_ids)
        ).order_by(models.Category.name).all()
    else:
        categories = db.query(models.Category).order_by(models.Category.name).all()

    all_tags = db.query(models.SkillTag).order_by(models.SkillTag.name).all()
    return templates.TemplateResponse(request, "skill_catalog.html", {
        "current_user": user, "skills": skills,
        "categories": categories,
        "sel_category": category_id, "sel_tier": tier,
        "all_tags": all_tags,
    })


@router.get("/skills/catalog/new", response_class=HTMLResponse)
def catalog_new_get(request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)
    categories = db.query(models.Category).order_by(models.Category.name).all()
    all_tags = db.query(models.SkillTag).order_by(models.SkillTag.name).all()
    return templates.TemplateResponse(request, "skill_catalog_form.html", {
        "current_user": user, "skill": None,
        "categories": categories, "error": None,
        "all_tags": all_tags,
    })


@router.post("/skills/catalog/new")
def catalog_new_post(
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    category_id: int = Form(0),
    tier: str = Form("basic"),
    tag_ids: List[int] = Form(default=[]),
    db: Session = Depends(get_db),
):
    user = auth.require_manager_or_admin(request, db)
    skill = models.Skill(
        name=name,
        description=description or None,
        category_id=category_id or None,
        tier=tier,
        created_by=user.id,
    )
    db.add(skill)
    db.flush()
    if tag_ids:
        skill.tags = db.query(models.SkillTag).filter(models.SkillTag.id.in_(tag_ids)).all()
    db.commit()
    return RedirectResponse("/skills/catalog", status_code=303)


@router.get("/skills/catalog/{skill_id}/edit", response_class=HTMLResponse)
def catalog_edit_get(skill_id: int, request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)
    skill = db.query(models.Skill).filter(models.Skill.id == skill_id).first()
    if not skill:
        return RedirectResponse("/skills/catalog", status_code=303)
    categories = db.query(models.Category).order_by(models.Category.name).all()
    all_tags = db.query(models.SkillTag).order_by(models.SkillTag.name).all()
    return templates.TemplateResponse(request, "skill_catalog_form.html", {
        "current_user": user, "skill": skill,
        "categories": categories, "error": None,
        "all_tags": all_tags,
    })


@router.post("/skills/catalog/{skill_id}/edit")
def catalog_edit_post(
    skill_id: int,
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    category_id: int = Form(0),
    tier: str = Form("basic"),
    tag_ids: List[int] = Form(default=[]),
    db: Session = Depends(get_db),
):
    user = auth.require_manager_or_admin(request, db)
    skill = db.query(models.Skill).filter(models.Skill.id == skill_id).first()
    if not skill:
        return RedirectResponse("/skills/catalog", status_code=303)
    skill.name = name
    skill.description = description or None
    skill.category_id = category_id or None
    skill.tier = tier
    # タグの更新
    if tag_ids:
        skill.tags = db.query(models.SkillTag).filter(models.SkillTag.id.in_(tag_ids)).all()
    else:
        skill.tags = []
    db.commit()
    return RedirectResponse("/skills/catalog", status_code=303)


@router.post("/skills/catalog/{skill_id}/delete")
def catalog_delete(skill_id: int, request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)
    skill = db.query(models.Skill).filter(models.Skill.id == skill_id).first()
    if skill:
        db.delete(skill)
        db.commit()
    return RedirectResponse("/skills/catalog", status_code=303)


# ────────────────────────────────────────────────────────────────
# インポート共通ヘルパー
# ────────────────────────────────────────────────────────────────
_VALID_TIERS = {"beginner", "basic", "intermediate", "advanced"}

_SAMPLE_RECORDS = [
    {"category": "HILS基盤・操作",  "color": "#dc2626", "name": "HILS基本操作",       "tier": "beginner",     "description": "HILSの電源投入・基本操作・ステータス確認"},
    {"category": "HILS基盤・操作",  "color": "#dc2626", "name": "HILSキャリブレーション","tier": "basic",      "description": "センサ・アクチュエータの校正・調整手順"},
    {"category": "dSPACEツール",    "color": "#2563eb", "name": "ControlDesk基本操作", "tier": "beginner",     "description": "レイアウト作成・変数モニタリング・データ記録"},
    {"category": "DevOps・自動化",  "color": "#059669", "name": "GitHub Actions",       "tier": "basic",        "description": "CI/CDパイプライン構築・自動テスト・通知連携"},
]


def _parse_csv_bytes(content: bytes) -> tuple[list, list]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp932", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    records, errors = [], []
    for i, row in enumerate(reader, start=2):
        name     = (row.get("skill_name") or row.get("name") or row.get("スキル名") or "").strip()
        category = (row.get("category_name") or row.get("category") or row.get("カテゴリ名") or row.get("カテゴリー") or "").strip()
        color    = (row.get("category_color") or row.get("color") or row.get("カラー") or "#f97316").strip()
        tier     = (row.get("tier") or row.get("ティア") or "basic").strip()
        desc     = (row.get("description") or row.get("説明") or "").strip()
        if not name:
            errors.append(f"行{i}: スキル名が空です"); continue
        if tier not in _VALID_TIERS:
            errors.append(f"行{i}: 無効なティア '{tier}'"); continue
        records.append({"category": category, "color": color, "name": name, "tier": tier, "description": desc})
    return records, errors


def _parse_excel_bytes(content: bytes) -> tuple[list, list]:
    import openpyxl
    import openpyxl.utils.exceptions
    records, errors = [], []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return [], [f"Excelファイルの読み込みに失敗しました: {e}"]

    # シート「スキル」か最初のシートを使用
    ws = wb["スキル"] if "スキル" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    def get(row, *keys):
        for k in keys:
            if k in col and row[col[k]].value is not None:
                return str(row[col[k]].value).strip()
        return ""

    # テンプレートは行1=英語ヘッダー・行2=日本語説明・行3以降=データ
    # 行2の日本語説明行をスキップするため min_row=3 から開始
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
        row = list(row)
        if not any(c.value for c in row):
            continue
        name     = get(row, "skill_name", "スキル名", "name")
        category = get(row, "category_name", "カテゴリ名", "category", "カテゴリー")
        color    = get(row, "category_color", "カラー", "color") or "#f97316"
        tier     = get(row, "tier", "ティア") or "basic"
        desc     = get(row, "description", "説明")
        if not name:
            errors.append(f"行{i}: スキル名が空です"); continue
        if tier not in _VALID_TIERS:
            errors.append(f"行{i}: 無効なティア '{tier}'"); continue
        records.append({"category": category, "color": color, "name": name, "tier": tier, "description": desc})
    return records, errors


def _parse_markdown_bytes(content: bytes) -> tuple[list, list]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp932", errors="replace")
    records, errors = [], []
    current_cat, current_color, current_tier = "", "#f97316", "basic"
    tier_map = {"beginner": "beginner", "ビギナー": "beginner", "初級": "beginner",
                "basic": "basic",       "ベーシック": "basic",   "基礎": "basic",
                "intermediate": "intermediate", "アドバンスド": "intermediate", "中級": "intermediate",
                "advanced": "advanced", "エキスパート": "advanced", "上級": "advanced"}
    for line_no, line in enumerate(text.splitlines(), start=1):
        line = line.rstrip()
        if line.startswith("## "):          # カテゴリ
            parts = line[3:].strip().split()
            current_color = parts[-1] if parts and parts[-1].startswith("#") else "#f97316"
            current_cat   = " ".join(p for p in parts if not p.startswith("#"))
        elif line.startswith("### "):       # ティア
            t = line[4:].strip()
            current_tier = tier_map.get(t, t.lower())
            if current_tier not in _VALID_TIERS:
                errors.append(f"行{line_no}: 不明なティア '{t}'（スキップ）"); current_tier = "basic"
        elif line.startswith("- "):         # スキル
            body = line[2:].strip()
            if ":" in body:
                name, desc = body.split(":", 1)
            else:
                name, desc = body, ""
            name = name.strip(); desc = desc.strip()
            if not name:
                errors.append(f"行{line_no}: スキル名が空です"); continue
            records.append({"category": current_cat, "color": current_color,
                            "name": name, "tier": current_tier, "description": desc})
    return records, errors


def _parse_json_bytes(content: bytes) -> tuple[list, list]:
    import json
    records, errors = [], []
    try:
        data = json.loads(content.decode("utf-8-sig"))
    except Exception as e:
        return [], [f"JSON解析エラー: {e}"]
    if not isinstance(data, list):
        return [], ["JSONはリスト形式（[...]）にしてください"]
    for ci, cat_obj in enumerate(data):
        cat_name  = str(cat_obj.get("category") or cat_obj.get("カテゴリ名") or "").strip()
        cat_color = str(cat_obj.get("color") or cat_obj.get("カラー") or "#f97316").strip()
        skills = cat_obj.get("skills") or []
        if not isinstance(skills, list):
            errors.append(f"カテゴリ[{ci}]: skills はリストにしてください"); continue
        for si, sk in enumerate(skills):
            name = str(sk.get("name") or sk.get("スキル名") or "").strip()
            tier = str(sk.get("tier") or sk.get("ティア") or "basic").strip()
            desc = str(sk.get("description") or sk.get("説明") or "").strip()
            if not name:
                errors.append(f"カテゴリ[{ci}]スキル[{si}]: スキル名が空です"); continue
            if tier not in _VALID_TIERS:
                errors.append(f"'{name}': 無効なティア '{tier}'"); continue
            records.append({"category": cat_name, "color": cat_color,
                            "name": name, "tier": tier, "description": desc})
    return records, errors


def _apply_records(records: list, user, db) -> tuple[int, int, list]:
    """records を DB に登録し (added, skipped, errors) を返す"""
    added = skipped = 0
    errors = []
    cat_cache: dict[str, models.Category] = {}

    for r in records:
        cat_name = r["category"]
        cat_obj  = None
        if cat_name:
            if cat_name in cat_cache:
                cat_obj = cat_cache[cat_name]
            else:
                cat_obj = db.query(models.Category).filter(models.Category.name == cat_name).first()
                if not cat_obj:
                    cat_obj = models.Category(name=cat_name, color=r["color"], created_by=user.id)
                    db.add(cat_obj); db.flush()
                cat_cache[cat_name] = cat_obj

        if db.query(models.Skill).filter(models.Skill.name == r["name"]).first():
            errors.append(f"'{r['name']}' は既に存在するためスキップ")
            skipped += 1
            continue

        db.add(models.Skill(
            name=r["name"],
            description=r["description"] or None,
            category_id=cat_obj.id if cat_obj else None,
            tier=r["tier"],
            created_by=user.id,
        ))
        added += 1

    db.commit()
    return added, skipped, errors


# ────────────────────────────────────────────────────────────────
# インポートエンドポイント
# ────────────────────────────────────────────────────────────────
@router.post("/skills/catalog/import")
async def catalog_import(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """CSV / Excel / Markdown / JSON からカテゴリ＋スキルを一括登録"""
    user = auth.require_manager_or_admin(request, db)
    content  = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".csv"):
        records, errors = _parse_csv_bytes(content)
    elif filename.endswith((".xlsx", ".xls")):
        records, errors = _parse_excel_bytes(content)
    elif filename.endswith((".md", ".markdown")):
        records, errors = _parse_markdown_bytes(content)
    elif filename.endswith(".json"):
        records, errors = _parse_json_bytes(content)
    else:
        return JSONResponse({"ok": False, "error": "対応形式: .csv .xlsx .md .json"}, status_code=400)

    added, skipped, apply_errors = _apply_records(records, user, db)
    errors += apply_errors
    return JSONResponse({"ok": True, "added": added, "skipped": skipped, "errors": errors})


# ────────────────────────────────────────────────────────────────
# スキルカタログ エクスポート（現在登録されているデータを出力）
# ────────────────────────────────────────────────────────────────
import json as _json

@router.get("/skills/catalog/export/{fmt}")
def catalog_export(fmt: str, request: Request, db: Session = Depends(get_db)):
    """スキルカタログを CSV / Excel / JSON / Markdown でエクスポート"""
    from fastapi.responses import StreamingResponse, Response
    user = auth.require_manager_or_admin(request, db)

    # Manager は自グループスコープ、Admin は全件
    if user.role == "manager":
        skill_ids = _get_manager_skill_ids(user, db)
        skills = db.query(models.Skill).filter(models.Skill.id.in_(skill_ids)).order_by(
            models.Skill.tier, models.Skill.name
        ).all() if skill_ids else []
    else:
        skills = db.query(models.Skill).order_by(models.Skill.tier, models.Skill.name).all()

    if fmt == "csv":
        lines = ["category_name,category_color,skill_name,tier,description"]
        for s in skills:
            cat_name  = (s.category.name  if s.category else "").replace(",", "、")
            cat_color = (s.category.color if s.category else "#f97316")
            desc = (s.description or "").replace(",", "、")
            lines.append(f'{cat_name},{cat_color},{s.name},{s.tier},{desc}')
        body = "﻿" + "\n".join(lines) + "\n"
        return Response(content=body.encode("utf-8"),
                        media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=skills.csv"})

    elif fmt == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "スキル"
        headers    = ["category_name", "category_color", "skill_name", "tier", "description"]
        headers_ja = ["カテゴリ名", "カラー(16進)", "スキル名", "ティア", "説明"]
        hfill = PatternFill("solid", fgColor="F97316")
        hfont = Font(bold=True, color="FFFFFF")
        for ci, (h, hj) in enumerate(zip(headers, headers_ja), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=ci, value=f"({hj})")
        widths = [24, 14, 30, 14, 40]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        for s in skills:
            ws.append([
                s.category.name  if s.category else "",
                s.category.color if s.category else "#f97316",
                s.name, s.tier, s.description or "",
            ])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf,
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=skills.xlsx"})

    elif fmt == "json":
        # カテゴリごとに階層化
        cat_map: dict[str, dict] = {}
        for s in skills:
            cat = s.category.name if s.category else "未分類"
            color = s.category.color if s.category else "#f97316"
            if cat not in cat_map:
                cat_map[cat] = {"category": cat, "color": color, "skills": []}
            cat_map[cat]["skills"].append({
                "name": s.name, "tier": s.tier, "description": s.description or ""
            })
        body = _json.dumps(list(cat_map.values()), ensure_ascii=False, indent=2)
        return Response(content=body.encode("utf-8"),
                        media_type="application/json; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=skills.json"})

    elif fmt == "markdown":
        lines = ["# スキルカタログ\n",
                 "<!-- 書式: ## カテゴリ名 #カラー → ### tier → - スキル名: 説明 -->", ""]
        current_cat = current_tier = None
        for s in skills:
            cat = s.category.name if s.category else "未分類"
            color = s.category.color if s.category else "#94a3b8"
            if cat != current_cat:
                current_cat = cat; current_tier = None
                lines += ["", f"## {cat} {color}", ""]
            if s.tier != current_tier:
                current_tier = s.tier
                lines += [f"### {s.tier}", ""]
            lines.append(f"- {s.name}: {s.description or ''}")
        lines.append("")
        body = "\n".join(lines)
        return Response(content=body.encode("utf-8"),
                        media_type="text/markdown; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=skills.md"})

    else:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail=f"不明な形式: {fmt}")


# ────────────────────────────────────────────────────────────────
# テンプレートダウンロード
# ────────────────────────────────────────────────────────────────

@router.get("/skills/catalog/template/{fmt}")
def catalog_template(fmt: str, request: Request, db: Session = Depends(get_db)):
    """インポート用テンプレートファイルをダウンロードする"""
    auth.require_manager_or_admin(request, db)

    from fastapi.responses import StreamingResponse, Response

    if fmt == "csv":
        lines = ["category_name,category_color,skill_name,tier,description"]
        for r in _SAMPLE_RECORDS:
            lines.append(f'{r["category"]},{r["color"]},{r["name"]},{r["tier"]},{r["description"]}')
        body = "﻿" + "\n".join(lines) + "\n"  # BOM付きUTF-8
        return Response(content=body.encode("utf-8"),
                        media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=skill_template.csv"})

    elif fmt == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "スキル"

        header_fill = PatternFill("solid", fgColor="F97316")
        header_font = Font(bold=True, color="FFFFFF")
        headers = ["category_name", "category_color", "skill_name", "tier", "description"]
        headers_ja = ["カテゴリ名", "カラー(16進)", "スキル名", "ティア", "説明"]

        for ci, (h_en, h_ja) in enumerate(zip(headers, headers_ja), start=1):
            cell = ws.cell(row=1, column=ci, value=h_en)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=ci, value=f"({h_ja})")

        col_widths = [22, 14, 28, 14, 40]
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        for ri, r in enumerate(_SAMPLE_RECORDS, start=3):
            ws.cell(row=ri, column=1, value=r["category"])
            ws.cell(row=ri, column=2, value=r["color"])
            ws.cell(row=ri, column=3, value=r["name"])
            ws.cell(row=ri, column=4, value=r["tier"])
            ws.cell(row=ri, column=5, value=r["description"])

        # 「説明」シートを追加
        ws2 = wb.create_sheet("使い方")
        notes = [
            ("項目", "説明"),
            ("category_name", "カテゴリ名。存在しない場合は新規作成されます。"),
            ("category_color", "カテゴリの色（16進数 例: #dc2626）。省略可。"),
            ("skill_name", "スキル名（必須・重複スキップ）"),
            ("tier", "難易度: beginner / basic / intermediate / advanced"),
            ("description", "スキルの説明（省略可）"),
        ]
        for row in notes:
            ws2.append(row)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf,
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=skill_template.xlsx"})

    elif fmt == "markdown":
        lines = ["# スキルカタログ インポートテンプレート", "",
                 "<!-- 書式: ## カテゴリ名 #カラー(16進) → ### ティア → - スキル名: 説明 -->", ""]
        current_cat = None
        current_tier = None
        for r in _SAMPLE_RECORDS:
            if r["category"] != current_cat:
                current_cat  = r["category"]
                current_tier = None
                lines += ["", f'## {r["category"]} {r["color"]}', ""]
            if r["tier"] != current_tier:
                current_tier = r["tier"]
                lines += [f'### {r["tier"]}', ""]
            lines.append(f'- {r["name"]}: {r["description"]}')
        lines.append("")
        body = "\n".join(lines)
        return Response(content=body.encode("utf-8"),
                        media_type="text/markdown; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=skill_template.md"})

    elif fmt == "json":
        # JSON は category ごとにまとめる
        cat_map: dict[str, dict] = {}
        for r in _SAMPLE_RECORDS:
            if r["category"] not in cat_map:
                cat_map[r["category"]] = {"category": r["category"], "color": r["color"], "skills": []}
            cat_map[r["category"]]["skills"].append(
                {"name": r["name"], "tier": r["tier"], "description": r["description"]}
            )
        body = _json.dumps(list(cat_map.values()), ensure_ascii=False, indent=2)
        return Response(content=body.encode("utf-8"),
                        media_type="application/json; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=skill_template.json"})

    else:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail=f"不明な形式: {fmt}")


# ════════════════════════════════════════════════════════════════
# ティア名カスタマイズ（Admin / Manager のみ）
# ════════════════════════════════════════════════════════════════

@router.get("/skills/tier-settings", response_class=HTMLResponse)
def tier_settings_get(request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)
    tier_names = models.get_tier_display_names(db)
    tier_order = ["beginner", "basic", "intermediate", "advanced"]
    return templates.TemplateResponse(request, "tier_settings.html", {
        "current_user": user,
        "tier_names": tier_names,
        "tier_order": tier_order,
        "TIER_ICONS": models.TIER_ICONS,
        "TIER_COLORS": models.TIER_COLORS,
        "success": False,
    })


@router.post("/skills/tier-settings")
def tier_settings_post(
    request: Request,
    tier_beginner: str = Form(""),
    tier_basic: str = Form(""),
    tier_intermediate: str = Form(""),
    tier_advanced: str = Form(""),
    db: Session = Depends(get_db),
):
    user = auth.require_manager_or_admin(request, db)
    mapping = {
        "beginner": tier_beginner.strip(),
        "basic": tier_basic.strip(),
        "intermediate": tier_intermediate.strip(),
        "advanced": tier_advanced.strip(),
    }
    for key, value in mapping.items():
        db_key = f"tier_name_{key}"
        setting = db.query(models.AppSetting).filter(
            models.AppSetting.key == db_key
        ).first()
        name = value or models.DEFAULT_TIER_NAMES[key]
        if setting:
            setting.value = name
        else:
            db.add(models.AppSetting(key=db_key, value=name))
    db.commit()

    tier_names = models.get_tier_display_names(db)
    tier_order = ["beginner", "basic", "intermediate", "advanced"]
    return templates.TemplateResponse(request, "tier_settings.html", {
        "current_user": user,
        "tier_names": tier_names,
        "tier_order": tier_order,
        "TIER_ICONS": models.TIER_ICONS,
        "TIER_COLORS": models.TIER_COLORS,
        "success": True,
    })


# ════════════════════════════════════════════════════════════════
# ユーザーのスキルレベル自己申告（全承認済みユーザー）
# ════════════════════════════════════════════════════════════════

@router.get("/skills", response_class=HTMLResponse)
def skills_my(
    request: Request,
    category_id: int = 0,
    tier: str = "",
    group_id: int = 0,
    view: str = "",
    db: Session = Depends(get_db),
):
    user = auth.require_approved(request, db)

    # カスタムティア名
    tier_names = models.get_tier_display_names(db)
    tier_order = ["beginner", "basic", "intermediate", "advanced"]

    # ユーザーの所属グループ
    my_groups = (
        db.query(models.Group)
        .join(models.GroupMembership)
        .filter(models.GroupMembership.user_id == user.id)
        .order_by(models.Group.name)
        .all()
    )

    # グループでスキル絞込み
    group_skill_ids = None
    if group_id:
        sel_group = db.query(models.Group).filter(models.Group.id == group_id).first()
        if sel_group:
            group_skill_ids = _get_all_group_skill_ids(sel_group)

    # 全カタログ取得（tierフィルタなし → 概要計算用）
    q_all = db.query(models.Skill)
    if category_id:
        q_all = q_all.filter(models.Skill.category_id == category_id)
    all_catalog = q_all.order_by(models.Skill.tier, models.Skill.name).all()
    if group_skill_ids is not None:
        all_catalog = [sk for sk in all_catalog if sk.id in group_skill_ids]

    # 自分のスキルレベル（全ステータス：承認状況表示用）
    my_levels = (
        db.query(models.UserSkillLevel)
        .filter(models.UserSkillLevel.user_id == user.id)
        .all()
    )
    # レベル表示・集計は承認済みのみ
    my_level_map: dict[int, int] = {
        sl.skill_id: sl.level for sl in my_levels if sl.approval_status == "approved"
    }
    # 申請中のレベル（ドット横に「申請中 〇〇」と表示するため）
    my_pending_level_map: dict[int, int] = {
        sl.skill_id: sl.level for sl in my_levels if sl.approval_status == "pending"
    }
    # ステータス列・ドット色は全ステータスを参照
    my_approval_map: dict[int, str] = {
        sl.skill_id: sl.approval_status for sl in my_levels
    }
    my_approver_map: dict[int, int] = {
        sl.skill_id: sl.approver_id for sl in my_levels if sl.approver_id
    }

    # ── ティア概要（overview_mode） ──
    overview_mode = not tier and view != "all"
    tier_summary = {}
    for t_key in tier_order:
        t_skills = [s for s in all_catalog if s.tier == t_key]
        t_acquired = sum(1 for s in t_skills if my_level_map.get(s.id, 0) > 0)
        tier_summary[t_key] = {"total": len(t_skills), "acquired": t_acquired}

    # tierが選択されている場合のみ詳細スキル一覧
    if tier:
        catalog = [sk for sk in all_catalog if sk.tier == tier]
    else:
        catalog = all_catalog

    categories = db.query(models.Category).order_by(models.Category.name).all()

    # 承認者候補: 自分が所属するグループの担当 Manager のみ
    # グループ未所属の場合は Manager/Admin 全員を fallback として表示
    from sqlalchemy import select as _select

    memberships = (
        db.query(models.GroupMembership)
        .filter(models.GroupMembership.user_id == user.id)
        .all()
    )
    if memberships:
        group_ids = [m.group_id for m in memberships]
        # group_managers テーブルから co-manager を ORM で取得（SQLite 互換）
        co_mgr_rows = db.execute(
            _select(models.group_managers.c.user_id)
            .where(models.group_managers.c.group_id.in_(group_ids))
            .distinct()
        ).fetchall()
        co_mgr_ids = {r[0] for r in co_mgr_rows}
        # primary manager_id も加える
        primary_mgr_ids = {
            g.manager_id
            for g in db.query(models.Group).filter(models.Group.id.in_(group_ids)).all()
            if g.manager_id
        }
        approver_ids = (co_mgr_ids | primary_mgr_ids) - {user.id}

        if approver_ids:
            approvers = (
                db.query(models.User)
                .filter(
                    models.User.id.in_(approver_ids),
                    models.User.is_approved == True,
                )
                .order_by(models.User.display_name, models.User.username)
                .all()
            )
        else:
            # 担当 Manager が見つからない場合は Manager/Admin 全員
            approvers = (
                db.query(models.User)
                .filter(
                    models.User.is_approved == True,
                    models.User.role.in_(["manager", "admin"]),
                    models.User.id != user.id,
                )
                .order_by(models.User.display_name, models.User.username)
                .all()
            )
    else:
        # グループ未所属: Manager/Admin 全員を fallback
        approvers = (
            db.query(models.User)
            .filter(
                models.User.is_approved == True,
                models.User.role.in_(["manager", "admin"]),
                models.User.id != user.id,
            )
            .order_by(models.User.display_name, models.User.username)
            .all()
        )

    by_tier: dict[str, list] = defaultdict(list)
    for sk in catalog:
        by_tier[sk.tier].append(sk)

    return templates.TemplateResponse(request, "skills.html", {
        "current_user": user,
        "by_tier": by_tier,
        "tier_order": tier_order,
        "my_level_map": my_level_map,
        "my_pending_level_map": my_pending_level_map,
        "my_approval_map": my_approval_map,
        "my_approver_map": my_approver_map,
        "approvers": approvers,
        "categories": categories,
        "sel_category": category_id,
        "sel_tier": tier,
        "total_catalog": len(all_catalog),
        "total_set": sum(1 for sk in all_catalog if my_level_map.get(sk.id, 0) > 0),
        "my_groups": my_groups,
        "sel_group": group_id,
        "overview_mode": overview_mode,
        "tier_summary": tier_summary,
        "TIER_NAMES": tier_names,
        "TIER_ICONS": models.TIER_ICONS,
        "TIER_DESCRIPTIONS": models.TIER_DESCRIPTIONS,
        "view": view,
    })


@router.post("/skills/{skill_id}/level")
def set_skill_level(
    skill_id: int,
    request: Request,
    level: int = Form(...),
    approver_id: int = Form(0),
    db: Session = Depends(get_db),
):
    """ユーザーが自分のスキルレベルを登録・更新する"""
    user = auth.require_approved(request, db)
    level = max(0, min(4, level))

    # Admin/Manager は自動承認
    is_auto_approve = user.role in ("admin", "manager")

    if not is_auto_approve:
        # 一般ユーザーは承認者必須
        approver = db.query(models.User).filter(
            models.User.id == approver_id,
            models.User.is_approved == True,
            models.User.id != user.id,
        ).first()
        if not approver:
            referer = request.headers.get("referer", "/skills")
            return RedirectResponse(referer, status_code=303)

    existing = (db.query(models.UserSkillLevel)
                .filter(
                    models.UserSkillLevel.user_id == user.id,
                    models.UserSkillLevel.skill_id == skill_id,
                ).first())

    if is_auto_approve:
        # 承認前のレベルを取得（履歴用）
        prev_history = (
            db.query(models.SkillLevelHistory)
            .filter(
                models.SkillLevelHistory.user_id == user.id,
                models.SkillLevelHistory.skill_id == skill_id,
            )
            .order_by(models.SkillLevelHistory.changed_at.desc())
            .first()
        )
        previous_level = prev_history.level if prev_history else 0

        if existing:
            existing.level = level
            existing.approver_id = None
            existing.approval_status = "approved"
            existing.approved_at = func.now()
            existing.approver_comment = "自動承認"
        else:
            db.add(models.UserSkillLevel(
                user_id=user.id, skill_id=skill_id, level=level,
                approver_id=None, approval_status="approved",
                approved_at=func.now(), approver_comment="自動承認",
            ))

        # 履歴を記録
        db.add(models.SkillLevelHistory(
            user_id=user.id,
            skill_id=skill_id,
            level=level,
            previous_level=previous_level,
            approved_by=user.id,
        ))
    else:
        if existing:
            existing.level = level
            existing.approver_id = approver_id
            existing.approval_status = "pending"
            existing.approved_at = None
            existing.approver_comment = None
        else:
            db.add(models.UserSkillLevel(
                user_id=user.id, skill_id=skill_id, level=level,
                approver_id=approver_id, approval_status="pending",
            ))

    db.commit()
    referer = request.headers.get("referer", "/skills")
    return RedirectResponse(referer, status_code=303)


# ════════════════════════════════════════════════════════════════
# 承認ワークフロー
# ════════════════════════════════════════════════════════════════

@router.get("/approvals", response_class=HTMLResponse)
def approvals_list(request: Request, db: Session = Depends(get_db)):
    """承認者として自分に割り当てられた承認依頼一覧（Admin/Managerのみ）"""
    user = auth.require_manager_or_admin(request, db)
    pending = (
        db.query(models.UserSkillLevel)
        .filter(
            models.UserSkillLevel.approver_id == user.id,
            models.UserSkillLevel.approval_status == "pending",
        )
        .all()
    )
    history = (
        db.query(models.UserSkillLevel)
        .filter(
            models.UserSkillLevel.approver_id == user.id,
            models.UserSkillLevel.approval_status.in_(["approved", "rejected"]),
        )
        .order_by(models.UserSkillLevel.approved_at.desc())
        .limit(50)
        .all()
    )
    return templates.TemplateResponse(request, "approvals.html", {
        "current_user": user,
        "pending": pending,
        "history": history,
    })


@router.post("/approvals/{record_id}/approve")
def approve_skill(
    record_id: int,
    request: Request,
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    """スキルレベルを承認する"""
    user = auth.require_manager_or_admin(request, db)
    record = db.query(models.UserSkillLevel).filter(
        models.UserSkillLevel.id == record_id,
        models.UserSkillLevel.approver_id == user.id,
        models.UserSkillLevel.approval_status == "pending",
    ).first()
    if record:
        # 承認前のレベルを取得（履歴用）
        prev_history = (
            db.query(models.SkillLevelHistory)
            .filter(
                models.SkillLevelHistory.user_id == record.user_id,
                models.SkillLevelHistory.skill_id == record.skill_id,
            )
            .order_by(models.SkillLevelHistory.changed_at.desc())
            .first()
        )
        previous_level = prev_history.level if prev_history else 0

        record.approval_status = "approved"
        record.approved_at = func.now()
        record.approver_comment = comment or None

        # 承認履歴を記録
        db.add(models.SkillLevelHistory(
            user_id=record.user_id,
            skill_id=record.skill_id,
            level=record.level,
            previous_level=previous_level,
            approved_by=user.id,
        ))
        db.commit()
    return RedirectResponse("/approvals", status_code=303)


@router.post("/approvals/{record_id}/reject")
def reject_skill(
    record_id: int,
    request: Request,
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    """スキルレベルを差し戻す"""
    user = auth.require_manager_or_admin(request, db)
    record = db.query(models.UserSkillLevel).filter(
        models.UserSkillLevel.id == record_id,
        models.UserSkillLevel.approver_id == user.id,
        models.UserSkillLevel.approval_status == "pending",
    ).first()
    if record:
        record.approval_status = "rejected"
        record.approved_at = func.now()
        record.approver_comment = comment or None
        db.commit()
    return RedirectResponse("/approvals", status_code=303)


@router.post("/api/approvals/bulk-action")
def bulk_approval_action(
    request: Request,
    action: str = Form(...),
    record_ids: List[int] = Form(default=[]),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    """一括承認・一括差し戻し（Admin/Managerのみ）"""
    user = auth.require_manager_or_admin(request, db)
    if action not in ("approve", "reject") or not record_ids:
        return JSONResponse({"ok": False, "error": "無効なリクエスト"}, status_code=400)

    records = (
        db.query(models.UserSkillLevel)
        .filter(
            models.UserSkillLevel.id.in_(record_ids),
            models.UserSkillLevel.approver_id == user.id,
            models.UserSkillLevel.approval_status == "pending",
        )
        .all()
    )

    processed = 0
    for record in records:
        if action == "approve":
            prev_history = (
                db.query(models.SkillLevelHistory)
                .filter(
                    models.SkillLevelHistory.user_id == record.user_id,
                    models.SkillLevelHistory.skill_id == record.skill_id,
                )
                .order_by(models.SkillLevelHistory.changed_at.desc())
                .first()
            )
            previous_level = prev_history.level if prev_history else 0
            record.approval_status = "approved"
            record.approved_at = func.now()
            record.approver_comment = comment or None
            db.add(models.SkillLevelHistory(
                user_id=record.user_id,
                skill_id=record.skill_id,
                level=record.level,
                previous_level=previous_level,
                approved_by=user.id,
            ))
        else:
            record.approval_status = "rejected"
            record.approved_at = func.now()
            record.approver_comment = comment or None
        processed += 1

    db.commit()
    return JSONResponse({"ok": True, "processed": processed})


@router.get("/approvals/my", response_class=HTMLResponse)
def my_approvals(request: Request, db: Session = Depends(get_db)):
    """自分が申請したスキルレベルの承認状況一覧"""
    user = auth.require_approved(request, db)
    records = (
        db.query(models.UserSkillLevel)
        .filter(models.UserSkillLevel.user_id == user.id)
        .order_by(models.UserSkillLevel.updated_at.desc())
        .all()
    )
    # 承認者リスト（再申請モーダル用）
    group_ids = [m.group_id for m in db.query(models.GroupMembership)
                 .filter(models.GroupMembership.user_id == user.id).all()]
    if group_ids:
        from sqlalchemy import select as sa_select
        co_mgr_ids = set(
            row[0] for row in db.execute(
                sa_select(models.group_managers.c.user_id).where(
                    models.group_managers.c.group_id.in_(group_ids)
                ).distinct()
            ).fetchall()
        )
        primary_mgr_ids = {
            g.manager_id
            for g in db.query(models.Group).filter(models.Group.id.in_(group_ids)).all()
            if g.manager_id
        }
        approver_ids = (co_mgr_ids | primary_mgr_ids) - {user.id}
        approvers = (
            db.query(models.User)
            .filter(models.User.id.in_(approver_ids), models.User.is_approved == True)
            .order_by(models.User.display_name, models.User.username)
            .all()
        ) if approver_ids else (
            db.query(models.User)
            .filter(models.User.is_approved == True,
                    models.User.role.in_(["manager", "admin"]),
                    models.User.id != user.id)
            .order_by(models.User.display_name, models.User.username)
            .all()
        )
    else:
        approvers = (
            db.query(models.User)
            .filter(models.User.is_approved == True,
                    models.User.role.in_(["manager", "admin"]),
                    models.User.id != user.id)
            .order_by(models.User.display_name, models.User.username)
            .all()
        )
    return templates.TemplateResponse(request, "my_approvals.html", {
        "current_user": user,
        "records": records,
        "approvers": approvers,
    })


# ════════════════════════════════════════════════════════════════
# JSON API（AJAX 用）
# ════════════════════════════════════════════════════════════════

@router.post("/api/skills/{skill_id}/level")
def api_set_skill_level(
    skill_id: int,
    request: Request,
    level: int = Form(...),
    approver_id: int = Form(0),
    db: Session = Depends(get_db),
):
    """AJAX: ユーザーのスキルレベルを登録・更新し JSON を返す"""
    user = auth.require_approved(request, db)
    level = max(0, min(4, level))

    is_auto_approve = user.role in ("admin", "manager")

    if not is_auto_approve:
        # 一般ユーザーは承認者必須
        approver = db.query(models.User).filter(
            models.User.id == approver_id,
            models.User.is_approved == True,
            models.User.id != user.id,
        ).first()
        if not approver:
            return JSONResponse({"ok": False, "error": "無効な承認者です"}, status_code=400)

    existing = (db.query(models.UserSkillLevel)
                .filter(
                    models.UserSkillLevel.user_id == user.id,
                    models.UserSkillLevel.skill_id == skill_id,
                ).first())

    if is_auto_approve:
        # 承認前のレベルを取得（履歴用）
        prev_history = (
            db.query(models.SkillLevelHistory)
            .filter(
                models.SkillLevelHistory.user_id == user.id,
                models.SkillLevelHistory.skill_id == skill_id,
            )
            .order_by(models.SkillLevelHistory.changed_at.desc())
            .first()
        )
        previous_level = prev_history.level if prev_history else 0

        if existing:
            existing.level = level
            existing.approver_id = None
            existing.approval_status = "approved"
            existing.approved_at = func.now()
            existing.approver_comment = "自動承認"
        else:
            db.add(models.UserSkillLevel(
                user_id=user.id, skill_id=skill_id, level=level,
                approver_id=None, approval_status="approved",
                approved_at=func.now(), approver_comment="自動承認",
            ))

        db.add(models.SkillLevelHistory(
            user_id=user.id,
            skill_id=skill_id,
            level=level,
            previous_level=previous_level,
            approved_by=user.id,
        ))
        db.commit()
        return JSONResponse({
            "ok": True,
            "skill_id": skill_id,
            "level": level,
            "level_name": models.SKILL_LEVELS[level],
            "level_color": models.LEVEL_COLORS[level],
            "approval_status": "approved",
            "approval_status_name": "承認済み",
        })
    else:
        if existing:
            existing.level = level
            existing.approver_id = approver_id
            existing.approval_status = "pending"
            existing.approved_at = None
            existing.approver_comment = None
        else:
            db.add(models.UserSkillLevel(
                user_id=user.id, skill_id=skill_id, level=level,
                approver_id=approver_id, approval_status="pending",
            ))
        db.commit()
        return JSONResponse({
            "ok": True,
            "skill_id": skill_id,
            "level": level,
            "level_name": models.SKILL_LEVELS[level],
            "level_color": models.LEVEL_COLORS[level],
            "approval_status": "pending",
            "approval_status_name": "承認待ち",
        })


@router.post("/api/approvals/my/{record_id}/withdraw")
def withdraw_my_approval(record_id: int, request: Request, db: Session = Depends(get_db)):
    """自分の承認待ち申請を取り下げる（レコード削除）"""
    user = auth.require_approved(request, db)
    rec = (
        db.query(models.UserSkillLevel)
        .filter(
            models.UserSkillLevel.id == record_id,
            models.UserSkillLevel.user_id == user.id,
            models.UserSkillLevel.approval_status == "pending",
        )
        .first()
    )
    if not rec:
        return JSONResponse({"ok": False, "error": "取り下げ可能な申請が見つかりません"}, status_code=404)
    db.delete(rec)
    db.commit()
    return JSONResponse({"ok": True})


@router.get("/api/dashboard/stats")
def api_dashboard_stats(request: Request, db: Session = Depends(get_db)):
    """AJAX: ダッシュボード用の集計 JSON を返す"""
    user = auth.require_approved(request, db)

    my_levels = (db.query(models.UserSkillLevel)
                 .filter(
                     models.UserSkillLevel.user_id == user.id,
                     models.UserSkillLevel.approval_status == "approved",
                 ).all())
    total = len(my_levels)
    catalog_total = db.query(models.Skill).count()
    avg_level = round(sum(sl.level for sl in my_levels) / total, 1) if total else 0.0

    level_dist = {str(i): 0 for i in range(5)}
    for sl in my_levels:
        level_dist[str(sl.level)] += 1

    cat_stats: dict[str, float] = {}
    cat_counts: dict[str, int] = {}
    for sl in my_levels:
        if sl.skill.category:
            n = sl.skill.category.name
            cat_stats[n] = cat_stats.get(n, 0.0) + sl.level
            cat_counts[n] = cat_counts.get(n, 0) + 1

    cat_avg = {}
    for n in cat_stats:
        cat_avg[n] = round(cat_stats[n] / cat_counts[n], 1) if cat_counts[n] else 0

    tier_stats = {}
    for tk in models.SKILL_TIERS:
        tier_total = db.query(models.Skill).filter(models.Skill.tier == tk).count()
        tier_done = sum(1 for sl in my_levels if sl.skill.tier == tk and sl.level > 0)
        tier_stats[tk] = {"total": tier_total, "done": tier_done}

    return JSONResponse({
        "total": total,
        "catalog_total": catalog_total,
        "avg_level": avg_level,
        "level_dist": level_dist,
        "cat_avg": cat_avg,
        "tier_stats": tier_stats,
    })


# ════════════════════════════════════════════════════════════════
# スキルマトリクス（社員×スキル ヒートマップ）
# ════════════════════════════════════════════════════════════════

@router.get("/skills/matrix", response_class=HTMLResponse)
def skill_matrix(
    request: Request,
    category_id: int = 0,
    group_id: int = 0,
    db: Session = Depends(get_db),
):
    """管理者/マネージャー向け: 社員のスキル状態を一覧表示"""
    user = auth.require_manager_or_admin(request, db)

    # Manager が閲覧可能なグループ ID
    is_manager = user.role == "manager"
    if is_manager:
        managed_group_ids = {g.id for g in user.managed_groups}
        managed_member_ids = {
            m.user_id
            for gid in managed_group_ids
            for m in db.query(models.GroupMembership)
                .filter(models.GroupMembership.group_id == gid).all()
        }

    # 対象ユーザーの絞り込み
    if group_id:
        # Manager は自分の担当グループのみ
        if is_manager and group_id not in managed_group_ids:
            group_id = 0
        group = db.query(models.Group).filter(models.Group.id == group_id).first() if group_id else None
        member_ids = [m.user_id for m in group.memberships] if group else []
        users = (db.query(models.User)
                 .filter(models.User.id.in_(member_ids),
                         models.User.role != "admin")
                 .order_by(models.User.display_name, models.User.username)
                 .all()) if member_ids else []
    else:
        if is_manager:
            # Manager: 担当グループのメンバーのみ（Admin 除外）
            users = (db.query(models.User)
                     .filter(models.User.id.in_(managed_member_ids),
                             models.User.is_approved == True,
                             models.User.role != "admin")
                     .order_by(models.User.display_name, models.User.username)
                     .all()) if managed_member_ids else []
        else:
            # Admin: 自分以外の全承認済みユーザー（Admin 除外）
            users = (db.query(models.User)
                     .filter(models.User.is_approved == True,
                             models.User.role != "admin")
                     .order_by(models.User.display_name, models.User.username)
                     .all())

    # スキルカタログ
    q = db.query(models.Skill)
    if category_id:
        q = q.filter(models.Skill.category_id == category_id)
    skills = q.order_by(models.Skill.category_id, models.Skill.name).all()

    # グループが選択されていて、そのグループにスキル割当がある場合はフィルタ（継承含む）
    if group_id:
        sel_group_obj = db.query(models.Group).filter(models.Group.id == group_id).first()
        if sel_group_obj:
            group_skill_ids = _get_all_group_skill_ids(sel_group_obj)
            if group_skill_ids:
                skills = [sk for sk in skills if sk.id in group_skill_ids]

    # 全承認済みスキルレベル取得
    user_ids = [u.id for u in users]
    all_levels = (
        db.query(models.UserSkillLevel)
        .filter(
            models.UserSkillLevel.user_id.in_(user_ids),
            models.UserSkillLevel.approval_status == "approved",
        )
        .all()
    ) if user_ids else []

    # {(user_id, skill_id): level} のマップ
    level_map = {}
    for sl in all_levels:
        level_map[(sl.user_id, sl.skill_id)] = sl.level

    categories = db.query(models.Category).order_by(models.Category.name).all()
    if is_manager:
        groups = (db.query(models.Group)
                  .filter(models.Group.id.in_(managed_group_ids))
                  .order_by(models.Group.name).all())
    else:
        groups = db.query(models.Group).order_by(models.Group.name).all()

    # ── 分析データ ──
    from datetime import timedelta
    from collections import defaultdict

    # 1) カテゴリー別平均（レーダーチャート用）
    cat_user_avg: dict[str, dict[str, float]] = {}  # {cat_name: {user_name: avg}}
    cat_names_ordered = []
    for cat in categories:
        cat_skills = [sk for sk in skills if sk.category_id == cat.id]
        if not cat_skills:
            continue
        cat_names_ordered.append(cat.name)
        cat_user_avg[cat.name] = {}
        for u in users:
            vals = [level_map.get((u.id, sk.id), 0) for sk in cat_skills]
            filled = [v for v in vals if v > 0]
            cat_user_avg[cat.name][u.display_name or u.username] = round(
                sum(filled) / len(filled), 2
            ) if filled else 0.0

    # 2) レベル分布（ドーナツ用）
    level_dist = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0}
    for u in users:
        for sk in skills:
            lv = level_map.get((u.id, sk.id), 0)
            level_dist[lv] += 1

    # 3) 成長トレンド（折れ線グラフ用）: 月別の平均レベル推移
    growth_trend: list[dict] = []
    if user_ids:
        history_all = (
            db.query(models.SkillLevelHistory)
            .filter(models.SkillLevelHistory.user_id.in_(user_ids))
            .order_by(models.SkillLevelHistory.changed_at.asc())
            .all()
        )
        monthly: dict[str, list[int]] = defaultdict(list)
        for rec in history_all:
            if rec.changed_at:
                month_key = (rec.changed_at + timedelta(hours=9)).strftime("%Y-%m")
                monthly[month_key].append(rec.level)
        for month_key in sorted(monthly.keys()):
            vals = monthly[month_key]
            growth_trend.append({
                "month": month_key,
                "avg": round(sum(vals) / len(vals), 2),
                "count": len(vals),
            })

    # 4) ユーザー別平均（横棒グラフ用）
    user_avg_ranking = []
    for u in users:
        vals = [level_map.get((u.id, sk.id), 0) for sk in skills]
        filled = [v for v in vals if v > 0]
        avg = round(sum(filled) / len(filled), 2) if filled else 0.0
        user_avg_ranking.append({
            "name": u.display_name or u.username,
            "avg": avg,
            "count": len(filled),
        })
    user_avg_ranking.sort(key=lambda x: x["avg"], reverse=True)

    return templates.TemplateResponse(request, "skill_matrix.html", {
        "current_user": user,
        "users": users,
        "skills": skills,
        "level_map": level_map,
        "categories": categories,
        "groups": groups,
        "sel_category": category_id,
        "sel_group": group_id,
        # 分析データ
        "cat_names_ordered": cat_names_ordered,
        "cat_user_avg": cat_user_avg,
        "level_dist": level_dist,
        "growth_trend": growth_trend,
        "user_avg_ranking": user_avg_ranking,
    })


# ════════════════════════════════════════════════════════════════
# メンバースキル詳細（Manager / Admin）
# ════════════════════════════════════════════════════════════════

@router.get("/members/{user_id}/skills", response_class=HTMLResponse)
def member_skill_detail(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Manager以上がメンバーの取得スキル状況を詳細確認"""
    current_user = auth.require_manager_or_admin(request, db)

    target = db.query(models.User).filter(models.User.id == user_id).first()
    if not target:
        return RedirectResponse("/dashboard", status_code=303)

    # Manager は自分の担当グループのメンバーのみ閲覧可
    if current_user.role == "manager":
        managed_group_ids = [g.id for g in current_user.managed_groups]
        is_member = (db.query(models.GroupMembership)
                     .filter(
                         models.GroupMembership.user_id == user_id,
                         models.GroupMembership.group_id.in_(managed_group_ids),
                     ).first())
        if not is_member and user_id != current_user.id:
            return RedirectResponse("/dashboard", status_code=303)

    # 承認済みスキルレベル
    skill_levels = (db.query(models.UserSkillLevel)
                    .filter(
                        models.UserSkillLevel.user_id == target.id,
                        models.UserSkillLevel.approval_status == "approved",
                    ).all())

    # カテゴリー別に分類
    categories = db.query(models.Category).order_by(models.Category.name).all()
    cat_skills: dict[str, list] = {}
    uncategorized = []
    for sl in skill_levels:
        if sl.level == 0:
            continue
        if sl.skill.category:
            cname = sl.skill.category.name
            if cname not in cat_skills:
                cat_skills[cname] = []
            cat_skills[cname].append(sl)
        else:
            uncategorized.append(sl)
    # カテゴリーごとにレベル降順
    for k in cat_skills:
        cat_skills[k].sort(key=lambda sl: sl.level, reverse=True)
    uncategorized.sort(key=lambda sl: sl.level, reverse=True)

    # サマリー
    total = sum(1 for sl in skill_levels if sl.level > 0)
    avg_level = round(sum(sl.level for sl in skill_levels if sl.level > 0) / total, 1) if total else 0.0
    catalog_total = db.query(models.Skill).count()

    # レベル分布
    level_dist = {i: 0 for i in range(5)}
    for sl in skill_levels:
        level_dist[sl.level] += 1

    # ティア別
    tier_stats = {}
    for tier_key in models.SKILL_TIERS:
        tier_catalog = db.query(models.Skill).filter(models.Skill.tier == tier_key).count()
        tier_done = sum(1 for sl in skill_levels if sl.skill.tier == tier_key and sl.level > 0)
        tier_stats[tier_key] = {"total": tier_catalog, "done": tier_done}

    # 所属グループ
    user_groups = (db.query(models.Group)
                   .join(models.GroupMembership)
                   .filter(models.GroupMembership.user_id == target.id)
                   .order_by(models.Group.name).all())

    # 成長履歴（直近）
    recent_history = (db.query(models.SkillLevelHistory)
                      .filter(models.SkillLevelHistory.user_id == target.id)
                      .order_by(models.SkillLevelHistory.changed_at.desc())
                      .limit(10).all())

    return templates.TemplateResponse(request, "member_detail.html", {
        "current_user": current_user,
        "target": target,
        "cat_skills": cat_skills,
        "uncategorized": uncategorized,
        "total": total,
        "avg_level": avg_level,
        "catalog_total": catalog_total,
        "level_dist": level_dist,
        "tier_stats": tier_stats,
        "user_groups": user_groups,
        "recent_history": recent_history,
        "categories": categories,
    })


# ════════════════════════════════════════════════════════════════
# スキル成長タイムライン
# ════════════════════════════════════════════════════════════════

@router.get("/skills/timeline", response_class=HTMLResponse)
def skill_timeline(
    request: Request,
    user_id: int = 0,
    db: Session = Depends(get_db),
):
    """スキルの成長を時系列で確認"""
    current_user = auth.require_approved(request, db)
    is_privileged = current_user.role in ("admin", "manager")

    # 表示対象ユーザー
    if user_id and is_privileged:
        target = db.query(models.User).filter(models.User.id == user_id).first()
        if not target:
            target = current_user
    else:
        target = current_user

    # ユーザー選択候補（管理者/マネージャー用）
    all_users = []
    if is_privileged:
        all_users = (db.query(models.User)
                     .filter(models.User.is_approved == True)
                     .order_by(models.User.display_name, models.User.username)
                     .all())

    # 成長履歴の取得
    history = (
        db.query(models.SkillLevelHistory)
        .filter(models.SkillLevelHistory.user_id == target.id)
        .order_by(models.SkillLevelHistory.changed_at.asc())
        .all()
    )

    # 最新の承認済みスキルレベル
    current_levels = (
        db.query(models.UserSkillLevel)
        .filter(
            models.UserSkillLevel.user_id == target.id,
            models.UserSkillLevel.approval_status == "approved",
        )
        .all()
    )

    # カテゴリーごとの平均レベル推移データを構築
    from datetime import timedelta
    cat_timeline: dict[str, list] = {}
    for rec in history:
        cat_name = rec.skill.category.name if rec.skill.category else "未分類"
        if cat_name not in cat_timeline:
            cat_timeline[cat_name] = []
        cat_timeline[cat_name].append({
            "date": (rec.changed_at + timedelta(hours=9)).strftime("%Y-%m-%d") if rec.changed_at else "",
            "skill": rec.skill.name,
            "level": rec.level,
            "prev": rec.previous_level or 0,
        })

    # スキルごとの成長推移データ
    skill_timeline: dict[str, list] = {}
    for rec in history:
        sname = rec.skill.name
        if sname not in skill_timeline:
            skill_timeline[sname] = []
        skill_timeline[sname].append({
            "date": (rec.changed_at + timedelta(hours=9)).strftime("%Y-%m-%d") if rec.changed_at else "",
            "level": rec.level,
        })

    # 成長サマリー
    growth_count = sum(1 for r in history if (r.previous_level or 0) < r.level)
    recent_history = list(reversed(history[-20:]))

    return templates.TemplateResponse(request, "skill_timeline.html", {
        "current_user": current_user,
        "target": target,
        "all_users": all_users,
        "history": recent_history,
        "current_levels": current_levels,
        "cat_timeline": cat_timeline,
        "skill_timeline": skill_timeline,
        "growth_count": growth_count,
        "total_changes": len(history),
        "is_privileged": is_privileged,
    })


@router.get("/api/skills/timeline/{target_user_id}")
def api_skill_timeline(
    target_user_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """AJAX: 特定ユーザーのスキル成長履歴JSON"""
    current_user = auth.require_approved(request, db)
    is_privileged = current_user.role in ("admin", "manager")
    if target_user_id != current_user.id and not is_privileged:
        return JSONResponse({"error": "権限がありません"}, status_code=403)

    history = (
        db.query(models.SkillLevelHistory)
        .filter(models.SkillLevelHistory.user_id == target_user_id)
        .order_by(models.SkillLevelHistory.changed_at.asc())
        .all()
    )

    from datetime import timedelta
    data = []
    for rec in history:
        data.append({
            "date": (rec.changed_at + timedelta(hours=9)).strftime("%Y-%m-%d") if rec.changed_at else "",
            "skill": rec.skill.name,
            "category": rec.skill.category.name if rec.skill.category else "未分類",
            "level": rec.level,
            "previous_level": rec.previous_level or 0,
        })

    return JSONResponse({"ok": True, "data": data})


# ════════════════════════════════════════════════════════════════
# スキルタグ管理
# ════════════════════════════════════════════════════════════════

@router.get("/tags", response_class=HTMLResponse)
def tags_list(request: Request, db: Session = Depends(get_db)):
    user = auth.require_approved(request, db)
    tags = db.query(models.SkillTag).order_by(models.SkillTag.name).all()
    try:
        return templates.TemplateResponse(request, "tags.html", {
            "current_user": user, "tags": tags
        })
    except Exception:
        return JSONResponse([
            {"id": t.id, "name": t.name, "color": t.color} for t in tags
        ])


@router.post("/tags/new")
def tag_new_post(
    request: Request,
    name: str = Form(...),
    color: str = Form("#6c757d"),
    db: Session = Depends(get_db),
):
    user = auth.require_manager_or_admin(request, db)
    if not db.query(models.SkillTag).filter(models.SkillTag.name == name).first():
        db.add(models.SkillTag(name=name, color=color))
        db.commit()
    return RedirectResponse("/tags", status_code=303)


@router.post("/tags/{tag_id}/delete")
def tag_delete(tag_id: int, request: Request, db: Session = Depends(get_db)):
    user = auth.require_manager_or_admin(request, db)
    tag = db.query(models.SkillTag).filter(models.SkillTag.id == tag_id).first()
    if tag:
        db.delete(tag)
        db.commit()
    return RedirectResponse("/tags", status_code=303)


# ════════════════════════════════════════════════════════════════
# スキル一括操作
# ════════════════════════════════════════════════════════════════

@router.post("/skills/bulk-action")
def skills_bulk_action(
    request: Request,
    skill_ids: List[int] = Form(default=[]),
    action: str = Form(default=""),
    category_id: Optional[int] = Form(default=None),
    tier: Optional[str] = Form(default=None),
    tag_id: Optional[int] = Form(default=None),
    db: Session = Depends(get_db),
):
    user = auth.require_manager_or_admin(request, db)
    if not skill_ids:
        return RedirectResponse("/skills/catalog", status_code=303)

    skills = db.query(models.Skill).filter(models.Skill.id.in_(skill_ids)).all()

    if action == "archive":
        for skill in skills:
            skill.is_archived = True
    elif action == "unarchive":
        for skill in skills:
            skill.is_archived = False
    elif action == "change_category" and category_id is not None:
        for skill in skills:
            skill.category_id = category_id or None
    elif action == "change_tier" and tier:
        for skill in skills:
            skill.tier = tier
    elif action == "add_tag" and tag_id is not None:
        tag = db.query(models.SkillTag).filter(models.SkillTag.id == tag_id).first()
        if tag:
            for skill in skills:
                if tag not in skill.tags:
                    skill.tags.append(tag)
    elif action == "delete":
        # スキルと関連する申告データを削除
        db.query(models.UserSkillLevel).filter(
            models.UserSkillLevel.skill_id.in_(skill_ids)
        ).delete(synchronize_session=False)
        db.query(models.SkillLevelHistory).filter(
            models.SkillLevelHistory.skill_id.in_(skill_ids)
        ).delete(synchronize_session=False)
        db.query(models.Skill).filter(
            models.Skill.id.in_(skill_ids)
        ).delete(synchronize_session=False)

    db.commit()
    return RedirectResponse("/skills/catalog", status_code=303)
